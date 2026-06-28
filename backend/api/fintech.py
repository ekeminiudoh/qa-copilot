"""Fintech QA test session — Moniepoint payment flow testing."""

import io
from datetime import datetime
from typing import List, Optional

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import Response
from pydantic import BaseModel
from sqlalchemy.ext.asyncio import AsyncSession

from backend.auth import get_current_user_db
from backend.db import get_db
from backend.db.models import UserDB

router = APIRouter(prefix="/api/fintech", tags=["fintech"])


class PaymentTestResult(BaseModel):
    test_type: str              # inbound | outbound | validation | reversal
    scenario: str
    amount: str = ""
    sender_account: str = ""
    receiver_account: str = ""
    balance_before: str = ""
    balance_after: str = ""
    debit_correct: str = "OK"
    credit_correct: str = "OK"
    balance_updated: str = "OK"
    transaction_recorded: str = "OK"
    notification_received: str = "OK"
    response_time_ok: str = "OK"
    issue_description: str = ""
    screenshot_url: str = ""


class FintechSessionRequest(BaseModel):
    session_name: str = ""
    environment: str = "Staging"    # Staging | UAT | Production
    currency: str = "NGN"
    tests: List[PaymentTestResult]


def _test_status(t: PaymentTestResult) -> str:
    checks = [
        t.debit_correct, t.credit_correct, t.balance_updated,
        t.transaction_recorded, t.notification_received, t.response_time_ok,
    ]
    return "Passed" if all(c == "OK" for c in checks) else "Failed"


@router.post("/report/excel")
async def generate_fintech_excel(
    request: FintechSessionRequest,
    current_user: UserDB = Depends(get_current_user_db),
    db: AsyncSession = Depends(get_db),
):
    """Generate Excel fintech test report for payment flow testing."""
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(status_code=501, detail="pip install openpyxl")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Fintech Test Results"

    GREEN  = PatternFill("solid", fgColor="00B050")
    RED    = PatternFill("solid", fgColor="FF4444")
    ORANGE = PatternFill("solid", fgColor="FFA500")
    BLUE   = PatternFill("solid", fgColor="003366")
    LGRAY  = PatternFill("solid", fgColor="F2F2F2")

    WHITE_BOLD = Font(bold=True, color="FFFFFF")
    BOLD       = Font(bold=True)
    thin       = Side(style="thin")
    border     = Border(left=thin, right=thin, top=thin, bottom=thin)
    center     = Alignment(horizontal="center", vertical="center")
    wrap       = Alignment(wrap_text=True, vertical="center")

    # Title
    ws.append([f"Fintech Payment Test Report — {request.session_name or 'QA Session'}"])
    ws.append([f"Date: {datetime.now().strftime('%Y-%m-%d %H:%M')}"])
    ws.append([f"Environment: {request.environment} | Currency: {request.currency}"])
    ws.append([f"Tested by: {current_user.username}"])
    ws.append([])
    ws["A1"].font = Font(bold=True, size=13, color="003366")

    headers = [
        "#", "Test Type", "Scenario", "Amount", "Sender Account", "Receiver Account",
        "Balance Before", "Balance After", "Debit ✓", "Credit ✓",
        "Balance Updated", "Transaction Recorded", "Notification", "Response Time",
        "Issues", "Screenshots", "Status",
    ]

    hdr_row = ws.max_row + 1
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=hdr_row, column=col)
        cell.fill = BLUE
        cell.font = WHITE_BOLD
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[hdr_row].height = 36

    # Result columns (col index → attribute)
    result_cols = {
        9: "debit_correct", 10: "credit_correct", 11: "balance_updated",
        12: "transaction_recorded", 13: "notification_received", 14: "response_time_ok",
    }

    type_colors = {
        "inbound":    "00B0F0",
        "outbound":   "7030A0",
        "validation": "FF9900",
        "reversal":   "FF0000",
    }

    for i, test in enumerate(request.tests, 1):
        status = _test_status(test)
        row_data = [
            i, test.test_type, test.scenario, test.amount,
            test.sender_account, test.receiver_account,
            test.balance_before, test.balance_after,
            test.debit_correct, test.credit_correct,
            test.balance_updated, test.transaction_recorded,
            test.notification_received, test.response_time_ok,
            test.issue_description, test.screenshot_url, status,
        ]
        ws.append(row_data)
        r = ws.max_row
        ws.row_dimensions[r].height = 18

        # All borders
        for col in range(1, len(headers) + 1):
            ws.cell(row=r, column=col).border = border

        # # centered
        ws.cell(row=r, column=1).alignment = center

        # Test type coloured
        tc = ws.cell(row=r, column=2)
        color = type_colors.get(test.test_type.lower(), "595959")
        tc.fill = PatternFill("solid", fgColor=color)
        tc.font = Font(bold=True, color="FFFFFF")
        tc.alignment = center

        # Result cells
        for col, attr in result_cols.items():
            cell = ws.cell(row=r, column=col)
            val = getattr(test, attr)
            cell.fill = GREEN if val == "OK" else RED
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = center

        # Issue
        ws.cell(row=r, column=15).alignment = wrap

        # Screenshot hyperlink
        sc = ws.cell(row=r, column=16)
        if test.screenshot_url:
            sc.hyperlink = test.screenshot_url
            sc.value = "View"
            sc.font = Font(color="0070C0", underline="single")

        # Status
        st_cell = ws.cell(row=r, column=17)
        st_cell.fill = GREEN if status == "Passed" else RED
        st_cell.font = Font(bold=True, color="FFFFFF")
        st_cell.alignment = center

        # Zebra stripe plain cols
        if i % 2 == 0:
            for col in [1, 3, 4, 5, 6, 7, 8, 15, 16]:
                cell = ws.cell(row=r, column=col)
                if not cell.fill or cell.fill.fgColor.rgb in ("00000000", "FFFFFFFF"):
                    cell.fill = LGRAY

    # Summary
    ws.append([])
    total  = len(request.tests)
    passed = sum(1 for t in request.tests if _test_status(t) == "Passed")
    failed = total - passed
    rate   = f"{passed/total*100:.1f}%" if total else "0%"
    ws.append(["", "SUMMARY", f"Total: {total}", f"Passed: {passed}",
               f"Failed: {failed}", "", "", "", "", "", "", "", "", "", f"Pass Rate: {rate}"])
    sr = ws.max_row
    ws.cell(row=sr, column=2).font = BOLD
    ws.cell(row=sr, column=4).fill = GREEN
    ws.cell(row=sr, column=4).font = Font(bold=True, color="FFFFFF")
    ws.cell(row=sr, column=5).fill = RED if failed else GREEN
    ws.cell(row=sr, column=5).font = Font(bold=True, color="FFFFFF")

    # Widths
    widths = [4, 13, 30, 12, 18, 18, 14, 14, 10, 10, 14, 18, 14, 14, 35, 12, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = f"A{hdr_row + 1}"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"fintech_test_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )
