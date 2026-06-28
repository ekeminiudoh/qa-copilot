"""Gaming QA test session — report generation and Playwright script generator."""

import io
from datetime import datetime
from typing import List

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import Response
from pydantic import BaseModel
from sqlalchemy.ext.asyncio import AsyncSession

from backend.auth import get_current_user_db
from backend.db import get_db
from backend.db.models import UserDB

router = APIRouter(prefix="/api/gaming", tags=["gaming"])


class GameResult(BaseModel):
    provider: str = ""
    game_name: str
    game_image: str = "OK"
    launch_and_play: str = "OK"
    debit: str = "OK"
    credit: str = "OK"
    min_max_bet: str = "OK"
    issue_description: str = ""
    screenshot_url: str = ""


class GamingSessionRequest(BaseModel):
    session_name: str = ""
    min_bet: str = "1"
    max_bet: str = "5000"
    currency: str = "LSL"
    games: List[GameResult]


def _game_status(game: GameResult) -> str:
    all_ok = all([
        game.game_image == "OK",
        game.launch_and_play == "OK",
        game.debit == "OK",
        game.credit == "OK",
        game.min_max_bet == "OK",
    ])
    return "Tested OK" if all_ok else "Open"


@router.post("/report/excel")
async def generate_gaming_excel(
    request: GamingSessionRequest,
    current_user: UserDB = Depends(get_current_user_db),
    db: AsyncSession = Depends(get_db),
):
    """Generate Excel gaming test report in standard QA format."""
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(status_code=501, detail="openpyxl not installed: pip install openpyxl")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Gaming Test Results"

    # ── Styles ────────────────────────────────────────────────────────────────
    GREEN  = PatternFill("solid", fgColor="00B050")
    RED    = PatternFill("solid", fgColor="FF4444")
    ORANGE = PatternFill("solid", fgColor="FFA500")
    BLUE   = PatternFill("solid", fgColor="1F4E79")
    LGRAY  = PatternFill("solid", fgColor="F2F2F2")

    WHITE_BOLD = Font(bold=True, color="FFFFFF")
    BOLD       = Font(bold=True)
    thin       = Side(style="thin")
    border     = Border(left=thin, right=thin, top=thin, bottom=thin)
    center     = Alignment(horizontal="center", vertical="center")
    wrap       = Alignment(wrap_text=True, vertical="center")

    # ── Title block ───────────────────────────────────────────────────────────
    ws.append([f"Gaming Test Report — {request.session_name or 'QA Session'}"])
    ws.append([f"Date: {datetime.now().strftime('%Y-%m-%d %H:%M')}"])
    ws.append([f"Min/Max Bet Requirement: {request.min_bet} / {request.max_bet} {request.currency}"])
    ws.append([f"Tested by: {current_user.username}"])
    ws.append([])
    ws["A1"].font = Font(bold=True, size=13, color="1F4E79")

    # ── Column headers ────────────────────────────────────────────────────────
    min_max_label = f"Min/Max Bet Limit ({request.min_bet}/{request.max_bet} {request.currency})"
    headers = [
        "#", "Provider", "Game List", "Game Image",
        "Launch and play", "Debit", "Credit",
        min_max_label, "Issues/Bug Descriptions", "Screenshots Links", "Status",
    ]
    hdr_row = ws.max_row + 1
    ws.append(headers)
    for col, _ in enumerate(headers, 1):
        cell = ws.cell(row=hdr_row, column=col)
        cell.fill   = BLUE
        cell.font   = WHITE_BOLD
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[hdr_row].height = 32

    # ── Data rows ─────────────────────────────────────────────────────────────
    result_cols = {4: "game_image", 5: "launch_and_play", 6: "debit", 7: "credit", 8: "min_max_bet"}
    prev_provider = ""

    for i, game in enumerate(request.games, 1):
        status = _game_status(game)
        provider_display = game.provider if game.provider != prev_provider else ""
        prev_provider = game.provider

        row_data = [
            i,
            provider_display,
            game.game_name,
            game.game_image,
            game.launch_and_play,
            game.debit,
            game.credit,
            game.min_max_bet,
            game.issue_description,
            game.screenshot_url,
            status,
        ]
        ws.append(row_data)
        r = ws.max_row
        ws.row_dimensions[r].height = 18

        # Zebra stripe
        row_fill = LGRAY if i % 2 == 0 else None

        for col in range(1, 12):
            cell = ws.cell(row=r, column=col)
            cell.border = border

        # # and provider
        ws.cell(row=r, column=1).alignment = center
        ws.cell(row=r, column=2).font = BOLD

        # Result cells — green/red
        for col, attr in result_cols.items():
            cell = ws.cell(row=r, column=col)
            val  = getattr(game, attr)
            cell.fill      = GREEN if val == "OK" else RED
            cell.font      = Font(bold=True, color="FFFFFF")
            cell.alignment = center

        # Issue description — wrap text
        ws.cell(row=r, column=9).alignment = wrap

        # Screenshot — hyperlink
        sc_cell = ws.cell(row=r, column=10)
        if game.screenshot_url:
            sc_cell.hyperlink = game.screenshot_url
            sc_cell.value     = "View Screenshot"
            sc_cell.font      = Font(color="0070C0", underline="single")

        # Status
        status_cell = ws.cell(row=r, column=11)
        status_cell.fill      = GREEN if status == "Tested OK" else ORANGE
        status_cell.font      = Font(bold=True, color="FFFFFF")
        status_cell.alignment = center

        if row_fill:
            for col in [1, 2, 3, 9, 10]:
                cell = ws.cell(row=r, column=col)
                if not cell.fill or cell.fill.fgColor.rgb in ("00000000", "FFFFFFFF"):
                    cell.fill = row_fill

    # ── Summary block ─────────────────────────────────────────────────────────
    ws.append([])
    total  = len(request.games)
    passed = sum(1 for g in request.games if _game_status(g) == "Tested OK")
    failed = total - passed
    rate   = f"{passed/total*100:.1f}%" if total else "0%"

    ws.append(["", "SUMMARY", f"Total Games: {total}",
               f"Passed: {passed}", f"Failed: {failed}",
               "", "", "", f"Pass Rate: {rate}"])
    sr = ws.max_row
    for col in [2, 3, 4, 5, 9]:
        ws.cell(row=sr, column=col).font = BOLD
    ws.cell(row=sr, column=4).fill = GREEN
    ws.cell(row=sr, column=4).font = Font(bold=True, color="FFFFFF")
    ws.cell(row=sr, column=5).fill = RED if failed else GREEN
    ws.cell(row=sr, column=5).font = Font(bold=True, color="FFFFFF")

    # ── Column widths ─────────────────────────────────────────────────────────
    widths = [4, 20, 28, 13, 16, 10, 10, 24, 38, 22, 13]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = f"A{hdr_row + 1}"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    fname = f"gaming_test_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )


@router.post("/playwright-script")
async def generate_playwright_script(
    request: GamingSessionRequest,
    current_user: UserDB = Depends(get_current_user_db),
):
    """Generate a Playwright automation script for gaming QA."""
    games_js = "\n".join(
        f'  {{ provider: "{g.provider}", name: {json_str(g.game_name)} }},'
        for g in request.games
    )

    script = f"""// QA Copilot — Auto-generated Gaming Test Script
// Session : {request.session_name or "Gaming QA"}
// Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")}
// Min/Max  : {request.min_bet}/{request.max_bet} {request.currency}

const {{ chromium }} = require("playwright");
const fs = require("fs");
const path = require("path");

const CONFIG = {{
  siteUrl:  "YOUR_GAMING_SITE_URL",   // ← replace
  username: "YOUR_USERNAME",           // ← replace
  password: "YOUR_PASSWORD",           // ← replace
  minBet:   "{request.min_bet}",
  maxBet:   "{request.max_bet}",
  currency: "{request.currency}",
}};

const GAMES = [
{games_js}
];

async function takeScreenshot(page, name) {{
  const dir = "screenshots";
  if (!fs.existsSync(dir)) fs.mkdirSync(dir);
  const file = path.join(dir, name.replace(/[^a-z0-9]/gi, "_") + ".png");
  await page.screenshot({{ path: file, fullPage: false }});
  return file;
}}

async function checkBlankScreen(page) {{
  return page.evaluate(() => {{
    const body = document.body;
    const bg = window.getComputedStyle(body).backgroundColor;
    const hasCanvas  = !!document.querySelector("canvas");
    const hasIframe  = !!document.querySelector("iframe");
    const hasContent = document.body.innerText.trim().length > 50;
    return !hasCanvas && !hasIframe && !hasContent;
  }});
}}

(async () => {{
  const browser = await chromium.launch({{ headless: false, slowMo: 300 }});
  const context = await browser.newContext({{ viewport: {{ width: 1280, height: 800 }} }});
  const page    = await context.newPage();
  const results = [];

  // ── LOGIN ─────────────────────────────────────────────────────────────────
  await page.goto(CONFIG.siteUrl);
  await page.waitForLoadState("networkidle");
  await page.fill("input[name='username'], #username, input[type='email']", CONFIG.username);
  await page.fill("input[name='password'], #password, input[type='password']", CONFIG.password);
  await page.click("button[type='submit'], #login-btn, .login-btn");
  await page.waitForLoadState("networkidle");
  console.log("✅ Logged in");

  // ── TEST EACH GAME ────────────────────────────────────────────────────────
  for (const game of GAMES) {{
    console.log(`\\nTesting: ${{game.provider}} — ${{game.name}}`);
    const result = {{
      provider: game.provider, name: game.name,
      game_image: "-", launch: "-", debit: "-", credit: "-", min_max: "-",
      issue: "", screenshot: "",
    }};

    try {{
      // Search for the game
      const searchBox = await page.$("input[placeholder*='search' i], [data-testid='search']");
      if (searchBox) {{
        await searchBox.fill(game.name);
        await page.waitForTimeout(1500);
      }}

      // Check thumbnail
      const img = await page.$(`img[alt*="${{game.name}}" i], .game-card img`);
      result.game_image = (img && await img.isVisible()) ? "OK" : "-";

      // Screenshot of thumbnail
      result.screenshot = await takeScreenshot(page, `${{game.provider}}_${{game.name}}_thumb`);

      // Launch game
      const card = await page.$(`[data-game-name="${{game.name}}"], .game-card:has-text("${{game.name}}")`);
      if (card) {{
        await card.click();
        await page.waitForTimeout(3000);

        const blank = await checkBlankScreen(page);
        if (blank) {{
          result.issue = "Game launch failed — blank white screen";
          result.screenshot = await takeScreenshot(page, `${{game.provider}}_${{game.name}}_FAIL`);
        }} else {{
          result.launch = "OK";
          result.screenshot = await takeScreenshot(page, `${{game.provider}}_${{game.name}}_launched`);
        }}
      }} else {{
        result.issue = "Game card not found on page";
      }}

    }} catch (err) {{
      result.issue = err.message.substring(0, 120);
      console.error("  Error:", err.message);
    }}

    results.push(result);
    const icon = result.launch === "OK" ? "✅" : "❌";
    console.log(`  ${{icon}} Image:${{result.game_image}} Launch:${{result.launch}} Issue:${{result.issue || "none"}}`);

    // Go back to game list
    await page.goBack().catch(() => page.goto(CONFIG.siteUrl));
    await page.waitForTimeout(1000);
  }}

  // ── SUMMARY ───────────────────────────────────────────────────────────────
  console.log("\\n═══════════════════════════════════");
  console.log("  GAMING TEST SUMMARY");
  console.log("═══════════════════════════════════");
  const passed = results.filter(r => r.launch === "OK").length;
  console.log(`  Total : ${{results.length}}`);
  console.log(`  Passed: ${{passed}}`);
  console.log(`  Failed: ${{results.length - passed}}`);
  console.log("───────────────────────────────────");
  results.forEach((r, i) => {{
    const s = r.launch === "OK" ? "✅ PASS" : "❌ FAIL";
    console.log(`  ${{i+1}}. [${{s}}] ${{r.provider}} — ${{r.name}}${{r.issue ? " | " + r.issue : ""}}`);
  }});

  fs.writeFileSync("gaming_results.json", JSON.stringify(results, null, 2));
  console.log("\\nResults saved to gaming_results.json");

  await browser.close();
}})();
"""
    return {"script": script, "filename": "gaming_test.js"}


def json_str(s: str) -> str:
    return '"' + s.replace('"', '\\"') + '"'
